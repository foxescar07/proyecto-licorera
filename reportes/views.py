from django.shortcuts import render
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse
from ventas.models import Venta, DetalleVenta
from productos.models import Producto
from inventario.models import Inventario
from .forms import FiltroReporteForm
import json
import zoneinfo
from django.core.serializers.json import DjangoJSONEncoder
from django.urls import reverse

# ReportLab
from reportlab.lib.pagesizes import letter, landscape  # type: ignore
from reportlab.lib import colors  # type: ignore
from reportlab.lib.units import cm  # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable  # type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT  # type: ignore
from io import BytesIO

# openpyxl (Excel)
from openpyxl import Workbook # type: ignore # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

ZONA_COLOMBIA = zoneinfo.ZoneInfo('America/Bogota')

# ─────────────────────────────────────────────
#  COLORES CORPORATIVOS CYS
# ─────────────────────────────────────────────
COLOR_PRIMARIO   = colors.HexColor('#1A2B3C')   # azul oscuro (PDF)
COLOR_ACENTO     = colors.HexColor('#4DA8DA')   # azul claro (PDF)
COLOR_VERDE      = colors.HexColor('#2ecc71')
COLOR_AMARILLO   = colors.HexColor('#f1c40f')
COLOR_ROJO       = colors.HexColor('#e74c3c')
COLOR_GRIS_FILA  = colors.HexColor('#F0F4F8')
COLOR_BLANCO     = colors.white
COLOR_TEXTO      = colors.HexColor('#1A2B3C')

# ─────────────────────────────────────────────
#  COLORES EXCEL — ROJO SUAVE CYS LICORERA
# ─────────────────────────────────────────────
XLS_ROJO_MARCA    = 'A3242D'   # rojo del logo, para el nombre "CYS Ltda."
XLS_ROJO_HEADER   = 'C0392B'   # cabecera de tablas (rojo medio)
XLS_ROJO_SUAVE    = 'FBE4E4'   # filas alternas (rosa muy suave)
XLS_ROJO_BORDE    = 'E8B4B4'   # bordes suaves
XLS_ROJO_KPI_BG   = 'FDF1F1'   # fondo de las tarjetas KPI
XLS_GRIS_TEXTO    = '4A4A4A'   # subtítulos / texto secundario
XLS_BLANCO        = 'FFFFFF'
XLS_VERDE_OK      = '2E8B57'   # valores positivos (entradas/ingresos)
XLS_ROJO_OUT      = 'C0392B'   # valores negativos (salidas)

XLS_FONT = 'Calibri'


# ─────────────────────────────────────────────
#  HELPER: construir PDF con encabezado estándar
# ─────────────────────────────────────────────
def _build_pdf(titulo, subtitulo, elementos, orientacion='portrait'):
    buffer = BytesIO()
    pagesize = landscape(letter) if orientacion == 'landscape' else letter
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagesize,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'Titulo', parent=styles['Title'],
        fontSize=18, textColor=COLOR_PRIMARIO,
        spaceAfter=2, alignment=TA_LEFT,
    )
    estilo_subtitulo = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#6B8CA0'),
        spaceAfter=8, alignment=TA_LEFT,
    )

    historia = []
    historia.append(Paragraph("CYS Ltda.", estilo_titulo))
    historia.append(Paragraph(titulo, ParagraphStyle(
        'T2', parent=styles['Heading1'],
        fontSize=13, textColor=COLOR_ACENTO, spaceAfter=2,
    )))
    historia.append(Paragraph(subtitulo, estilo_subtitulo))
    historia.append(HRFlowable(width="100%", thickness=1.5, color=COLOR_ACENTO, spaceAfter=10))
    historia.extend(elementos)

    doc.build(historia)
    buffer.seek(0)
    return buffer


def _estilo_tabla_base(num_cols, col_cabecera=True):
    """Devuelve un TableStyle estándar CYS."""
    estilo = [
        # Cabecera
        ('BACKGROUND',    (0, 0), (-1, 0), COLOR_PRIMARIO),
        ('TEXTCOLOR',     (0, 0), (-1, 0), COLOR_BLANCO),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 9),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('TOPPADDING',    (0, 0), (-1, 0), 7),
        # Cuerpo
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('ALIGN',         (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [COLOR_BLANCO, COLOR_GRIS_FILA]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#D0DCE8')),
        ('BOX',           (0, 0), (-1, -1), 1,   COLOR_ACENTO),
    ]
    return TableStyle(estilo)


def _kpi_table(items):
    """
    items = [('Label', 'Valor', color_hex_str), ...]
    Devuelve una tabla de KPIs horizontal.
    """
    styles = getSampleStyleSheet()
    filas_label = []
    filas_val   = []
    for label, valor, color_hex in items:
        filas_label.append(Paragraph(
            f'<font color="{color_hex}"><b>{valor}</b></font>',
            ParagraphStyle('kv', fontSize=14, alignment=TA_CENTER)
        ))
        filas_val.append(Paragraph(
            label,
            ParagraphStyle('kl', fontSize=7, textColor=colors.HexColor('#6B8CA0'), alignment=TA_CENTER)
        ))

    data = [filas_label, filas_val]
    col_w = [4*cm] * len(items)
    t = Table(data, colWidths=col_w)
    t.setStyle(TableStyle([
        ('BOX',           (0, 0), (-1, -1), 1,   colors.HexColor('#D0DCE8')),
        ('INNERGRID',     (0, 0), (-1, -1), 0.4, colors.HexColor('#D0DCE8')),
        ('BACKGROUND',    (0, 0), (-1, -1), COLOR_GRIS_FILA),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    return t


# ═══════════════════════════════════════════════════════
#  GENERADORES PDF POR TIPO
# ═══════════════════════════════════════════════════════

def _pdf_ventas(ventas_qs, fecha_inicio, fecha_fin):
    styles = getSampleStyleSheet()
    elementos = []

    # KPIs
    total_v = sum(v.total_venta for v in ventas_qs)
    total_u = sum(det.cantidad for v in ventas_qs for det in v.detalles.all())
    total_c = ventas_qs.values('cliente').distinct().count()
    elementos.append(_kpi_table([
        ('Total Ventas',    f'${total_v:,.0f}', '#2ecc71'),
        ('Unidades',        str(total_u),        '#4DA8DA'),
        ('Clientes únicos', str(total_c),        '#f1c40f'),
    ]))
    elementos.append(Spacer(1, 12))

    # Tabla principal
    cabecera = [['#', 'Fecha', 'Cliente', 'Producto', 'Cant.', 'Precio Unit.', 'Total']]
    filas = cabecera
    contador = 1
    for v in ventas_qs:
        for det in v.detalles.all():
            filas.append([
                str(contador),
                v.fecha.astimezone(ZONA_COLOMBIA).strftime("%Y-%m-%d"),
                str(v.cliente),
                det.producto.nombre,
                str(det.cantidad),
                f'${float(det.precio_unitario):,.0f}',
                f'${float(det.subtotal()):,.0f}',
            ])
            contador += 1

    if len(filas) == 1:
        filas.append(['', 'Sin ventas en el período seleccionado', '', '', '', '', ''])

    col_w = [1*cm, 2.5*cm, 4*cm, 5*cm, 1.5*cm, 3*cm, 3*cm]
    t = Table(filas, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabla_base(7)
    estilo.add('ALIGN', (4, 1), (6, -1), 'RIGHT')
    t.setStyle(estilo)
    elementos.append(t)

    subtitulo = f"Período: {fecha_inicio or 'Todo'} → {fecha_fin or 'Todo'}  |  Generado: {timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M')}"
    return _build_pdf("Historial de Ventas", subtitulo, elementos, orientacion='landscape')


def _pdf_inventario(productos_qs, entradas_qs, salidas_qs):
    elementos = []
    styles = getSampleStyleSheet()

    total_reg   = productos_qs.count()
    en_stock    = sum(1 for p in productos_qs if p.stock_total > 10)
    stock_bajo  = sum(1 for p in productos_qs if 0 < p.stock_total <= 10)
    agotados    = sum(1 for p in productos_qs if p.stock_total == 0)

    elementos.append(_kpi_table([
        ('Registrados', str(total_reg),  '#4DA8DA'),
        ('En stock',    str(en_stock),   '#2ecc71'),
        ('Stock bajo',  str(stock_bajo), '#f1c40f'),
        ('Agotados',    str(agotados),   '#e74c3c'),
    ]))
    elementos.append(Spacer(1, 12))

    elementos.append(Paragraph("Estado de Stock", ParagraphStyle(
        'sh', fontSize=10, textColor=COLOR_PRIMARIO, fontName='Helvetica-Bold', spaceAfter=4)))

    filas = [['#', 'Producto', 'Stock', 'Estado']]
    for i, p in enumerate(productos_qs, 1):
        qty = p.stock_total
        if qty == 0:
            estado, color = 'Agotado',    COLOR_ROJO
        elif qty <= 10:
            estado, color = 'Stock bajo', COLOR_AMARILLO
        else:
            estado, color = 'En stock',   COLOR_VERDE
        filas.append([
            str(i), p.nombre, str(qty),
            Paragraph(f'<font color="{color.hexval()}"><b>{estado}</b></font>',
                      ParagraphStyle('e', fontSize=8)),
        ])

    col_w = [1*cm, 9*cm, 2.5*cm, 3*cm]
    t = Table(filas, colWidths=col_w, repeatRows=1)
    t.setStyle(_estilo_tabla_base(4))
    elementos.append(t)
    elementos.append(Spacer(1, 14))

    elementos.append(Paragraph("Entradas de Inventario", ParagraphStyle(
        'sh2', fontSize=10, textColor=COLOR_VERDE, fontName='Helvetica-Bold', spaceAfter=4)))

    filas_e = [['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']]
    for i, e in enumerate(entradas_qs, 1):
        filas_e.append([
            str(i), e.presentacion.producto.nombre, f'+{e.cantidad}',
            e.motivo or '—',
            e.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M"),
        ])
    if len(filas_e) == 1:
        filas_e.append(['', 'Sin entradas registradas', '', '', ''])

    col_w_e = [1*cm, 7*cm, 2.5*cm, 4*cm, 3.5*cm]
    te = Table(filas_e, colWidths=col_w_e, repeatRows=1)
    estilo_e = _estilo_tabla_base(5)
    estilo_e.add('TEXTCOLOR', (2, 1), (2, -1), COLOR_VERDE)
    estilo_e.add('FONTNAME',  (2, 1), (2, -1), 'Helvetica-Bold')
    te.setStyle(estilo_e)
    elementos.append(te)
    elementos.append(Spacer(1, 14))

    elementos.append(Paragraph("Salidas de Inventario", ParagraphStyle(
        'sh3', fontSize=10, textColor=COLOR_ROJO, fontName='Helvetica-Bold', spaceAfter=4)))

    filas_s = [['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']]
    for i, s in enumerate(salidas_qs, 1):
        filas_s.append([
            str(i), s.presentacion.producto.nombre, f'-{s.cantidad}',
            s.motivo or '—',
            s.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M"),
        ])
    if len(filas_s) == 1:
        filas_s.append(['', 'Sin salidas registradas', '', '', ''])

    col_w_s = [1*cm, 7*cm, 2.5*cm, 4*cm, 3.5*cm]
    ts = Table(filas_s, colWidths=col_w_s, repeatRows=1)
    estilo_s = _estilo_tabla_base(5)
    estilo_s.add('TEXTCOLOR', (2, 1), (2, -1), COLOR_ROJO)
    estilo_s.add('FONTNAME',  (2, 1), (2, -1), 'Helvetica-Bold')
    ts.setStyle(estilo_s)
    elementos.append(ts)

    subtitulo = f"Generado: {timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M')}"
    return _build_pdf("Reporte de Inventario", subtitulo, elementos)


def _pdf_proveedores(proveedores_list):
    elementos = []

    elementos.append(_kpi_table([
        ('Total Proveedores', str(len(proveedores_list)), '#5DCAA5'),
        ('Estado',            'Activos',                  '#4DA8DA'),
    ]))
    elementos.append(Spacer(1, 12))

    filas = [['#', 'Contacto', 'Empresa', 'Teléfono', 'Correo']]
    for i, p in enumerate(proveedores_list, 1):
        filas.append([
            str(i),
            p['nombre_contacto'],
            p['nombre_empresa'],
            p['telefono'],
            p['email'],
        ])

    col_w = [1*cm, 4*cm, 5.5*cm, 3.5*cm, 5.5*cm]
    t = Table(filas, colWidths=col_w, repeatRows=1)
    t.setStyle(_estilo_tabla_base(5))
    elementos.append(t)

    subtitulo = f"Generado: {timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M')}"
    return _build_pdf("Reporte de Proveedores", subtitulo, elementos)


def _pdf_resumen_diario(hoy, ventas_hoy, entradas_hoy, salidas_hoy,
                        ingresos_hoy, total_entradas_hoy, total_salidas_hoy,
                        top_productos_hoy):
    elementos = []
    styles = getSampleStyleSheet()

    elementos.append(_kpi_table([
        ('Ingresos hoy',  f'${ingresos_hoy:,.0f}',    '#2ecc71'),
        ('Ventas hoy',    str(len(ventas_hoy)),         '#4DA8DA'),
        ('Und. Entrada',  str(total_entradas_hoy),      '#f1c40f'),
        ('Und. Salida',   str(total_salidas_hoy),       '#e74c3c'),
    ]))
    elementos.append(Spacer(1, 12))

    def seccion(titulo, color_hex):
        elementos.append(Paragraph(titulo, ParagraphStyle(
            'sec', fontSize=10, textColor=colors.HexColor(color_hex),
            fontName='Helvetica-Bold', spaceAfter=4, spaceBefore=10)))

    seccion("Ventas del día", '#4DA8DA')
    filas_v = [['#', 'Cliente', 'Producto', 'Cant.', 'Total', 'Hora']]
    cont = 1
    for v in ventas_hoy:
        for det in v.detalles.all():
            filas_v.append([
                str(cont), str(v.cliente), det.producto.nombre,
                str(det.cantidad), f'${float(det.subtotal()):,.0f}',
                v.fecha.astimezone(ZONA_COLOMBIA).strftime("%H:%M"),
            ])
            cont += 1
    if len(filas_v) == 1:
        filas_v.append(['', 'Sin ventas hoy', '', '', '', ''])

    col_v = [1*cm, 4.5*cm, 5*cm, 1.5*cm, 3*cm, 2*cm]
    tv = Table(filas_v, colWidths=col_v, repeatRows=1)
    estilo_v = _estilo_tabla_base(6)
    estilo_v.add('ALIGN', (4, 1), (4, -1), 'RIGHT')
    tv.setStyle(estilo_v)
    elementos.append(tv)

    seccion("Entradas del día", '#2ecc71')
    filas_e = [['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']]
    for i, e in enumerate(entradas_hoy, 1):
        filas_e.append([str(i), e.presentacion.producto.nombre, f'+{e.cantidad}',
                         e.motivo or '—',
                         e.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M")])
    if len(filas_e) == 1:
        filas_e.append(['', 'Sin entradas hoy', '', '', ''])
    col_e = [1*cm, 6*cm, 2.5*cm, 4*cm, 3.5*cm]
    te = Table(filas_e, colWidths=col_e, repeatRows=1)
    te.setStyle(_estilo_tabla_base(5))
    elementos.append(te)

    seccion("Salidas del día", '#e74c3c')
    filas_s = [['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']]
    for i, s in enumerate(salidas_hoy, 1):
        filas_s.append([str(i), s.presentacion.producto.nombre, f'-{s.cantidad}',
                         s.motivo or '—',
                         s.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M")])
    if len(filas_s) == 1:
        filas_s.append(['', 'Sin salidas hoy', '', '', ''])
    col_s = [1*cm, 6*cm, 2.5*cm, 4*cm, 3.5*cm]
    ts = Table(filas_s, colWidths=col_s, repeatRows=1)
    ts.setStyle(_estilo_tabla_base(5))
    elementos.append(ts)

    if top_productos_hoy:
        seccion("Top productos del día", '#7F77DD')
        filas_t = [['#', 'Producto', 'Unidades vendidas', 'Total generado']]
        for i, (nombre, datos) in enumerate(top_productos_hoy, 1):
            filas_t.append([str(i), nombre, str(datos['cantidad']),
                             f'${datos["subtotal"]:,.0f}'])
        col_t = [1*cm, 8*cm, 4*cm, 4*cm]
        tt = Table(filas_t, colWidths=col_t, repeatRows=1)
        tt.setStyle(_estilo_tabla_base(4))
        elementos.append(tt)

    subtitulo = f"Fecha: {hoy.strftime('%d/%m/%Y')}  |  Generado: {timezone.now().astimezone(ZONA_COLOMBIA).strftime('%H:%M')}"
    return _build_pdf("Resumen Diario", subtitulo, elementos)


def _pdf_analisis_ventas(ventas_qs, total_ventas, total_productos, total_clientes, top_productos_hoy):
    elementos = []
    styles = getSampleStyleSheet()

    elementos.append(_kpi_table([
        ('Ventas Totales', f'${total_ventas:,.0f}', '#2ecc71'),
        ('Uds. Vendidas',  str(total_productos),    '#4DA8DA'),
        ('Clientes',       str(total_clientes),     '#f1c40f'),
    ]))
    elementos.append(Spacer(1, 12))

    elementos.append(Paragraph("Ventas por fecha", ParagraphStyle(
        'av1', fontSize=10, textColor=colors.HexColor('#E8834A'),
        fontName='Helvetica-Bold', spaceAfter=4)))

    filas = [['Fecha', 'Cliente', 'Producto', 'Cantidad', 'Total']]
    for v in ventas_qs:
        for det in v.detalles.all():
            filas.append([
                v.fecha.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y"),
                str(v.cliente),
                det.producto.nombre,
                str(det.cantidad),
                f'${float(det.subtotal()):,.0f}',
            ])
    if len(filas) == 1:
        filas.append(['Sin ventas registradas', '', '', '', ''])

    col_w = [2.5*cm, 4.5*cm, 6*cm, 2*cm, 3*cm]
    t = Table(filas, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabla_base(5)
    estilo.add('ALIGN', (4, 1), (4, -1), 'RIGHT')
    t.setStyle(estilo)
    elementos.append(t)
    elementos.append(Spacer(1, 14))

    elementos.append(Paragraph("Top productos del día más vendidos", ParagraphStyle(
        'av2', fontSize=10, textColor=colors.HexColor('#5DCAA5'),
        fontName='Helvetica-Bold', spaceAfter=4)))

    filas_top = [['#', 'Producto', 'Unidades vendidas', 'Total generado']]
    for i, (nombre, datos) in enumerate(top_productos_hoy, 1):
        filas_top.append([str(i), nombre, str(datos['cantidad']),
                           f'${datos["subtotal"]:,.0f}'])
    if len(filas_top) == 1:
        filas_top.append(['', 'Sin datos de productos hoy', '', ''])

    col_top = [1*cm, 9*cm, 4*cm, 4*cm]
    tt = Table(filas_top, colWidths=col_top, repeatRows=1)
    tt.setStyle(_estilo_tabla_base(4))
    elementos.append(tt)

    subtitulo = f"Generado: {timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M')}"
    return _build_pdf("Análisis de Ventas", subtitulo, elementos, orientacion='landscape')


# ═══════════════════════════════════════════════════════
#  GENERADORES EXCEL POR TIPO — TEMA "ROJO SUAVE" CYS
# ═══════════════════════════════════════════════════════

def _xls_fills():
    return {
        'marca':  PatternFill('solid', fgColor=XLS_ROJO_MARCA),
        'header': PatternFill('solid', fgColor=XLS_ROJO_HEADER),
        'suave':  PatternFill('solid', fgColor=XLS_ROJO_SUAVE),
        'blanco': PatternFill('solid', fgColor=XLS_BLANCO),
        'kpi_bg': PatternFill('solid', fgColor=XLS_ROJO_KPI_BG),
    }


def _xls_borde_suave():
    lado = Side(style='thin', color=XLS_ROJO_BORDE)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _xls_encabezado(ws, titulo, subtitulo, meta_lineas, num_cols):
    """Escribe el bloque 'CYS Ltda.' + título + subtítulo + metadatos.
    Devuelve la fila donde debe empezar la siguiente tabla."""
    ultima_col = get_column_letter(max(num_cols, 4))

    ws.merge_cells(f'A1:{ultima_col}1')
    c = ws['A1']
    c.value = 'CYS Ltda.'
    c.font = Font(name=XLS_FONT, size=18, bold=True, color=XLS_ROJO_MARCA)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{ultima_col}2')
    c = ws['A2']
    c.value = titulo
    c.font = Font(name=XLS_FONT, size=13, bold=True, color=XLS_ROJO_HEADER)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(f'A3:{ultima_col}3')
    c = ws['A3']
    c.value = subtitulo
    c.font = Font(name=XLS_FONT, size=9, italic=True, color=XLS_GRIS_TEXTO)
    ws.row_dimensions[3].height = 16

    fila = 4
    if meta_lineas:
        for etiqueta, valor in meta_lineas:
            ws.merge_cells(f'A{fila}:{get_column_letter(2)}{fila}')
            ws.merge_cells(f'{get_column_letter(3)}{fila}:{ultima_col}{fila}')
            ce = ws.cell(row=fila, column=1, value=etiqueta)
            ce.font = Font(name=XLS_FONT, size=9, bold=True, color=XLS_GRIS_TEXTO)
            cv = ws.cell(row=fila, column=3, value=valor)
            cv.font = Font(name=XLS_FONT, size=9, color=XLS_GRIS_TEXTO)
            fila += 1

    return fila + 1  # deja una fila en blanco antes de la siguiente tabla


def _xls_kpis(ws, fila, kpis, num_cols):
    """kpis = [(label, valor, color_hex)]. Se dibujan como tarjetas en una fila."""
    fills = _xls_fills()
    borde = _xls_borde_suave()
    ultima_col = max(num_cols, len(kpis))
    ancho_por_kpi = max(1, ultima_col // len(kpis))

    col = 1
    for label, valor, color_hex in kpis:
        fin = min(col + ancho_por_kpi - 1, ultima_col)
        if fin > col:
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=fin)
            ws.merge_cells(start_row=fila + 1, start_column=col, end_row=fila + 1, end_column=fin)
        cv = ws.cell(row=fila, column=col, value=valor)
        cv.font = Font(name=XLS_FONT, size=15, bold=True, color=color_hex)
        cv.alignment = Alignment(horizontal='center', vertical='center')
        cv.fill = fills['kpi_bg']
        cl = ws.cell(row=fila + 1, column=col, value=label.upper())
        cl.font = Font(name=XLS_FONT, size=8, bold=True, color=XLS_GRIS_TEXTO)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.fill = fills['kpi_bg']
        for r in (fila, fila + 1):
            for cc in range(col, fin + 1):
                ws.cell(row=r, column=cc).border = borde
        col = fin + 1

    ws.row_dimensions[fila].height = 26
    ws.row_dimensions[fila + 1].height = 16
    return fila + 3  # deja una fila en blanco


def _xls_seccion(ws, fila, texto, num_cols):
    ultima_col = get_column_letter(max(num_cols, 4))
    ws.merge_cells(f'A{fila}:{ultima_col}{fila}')
    c = ws.cell(row=fila, column=1, value=texto.upper())
    c.font = Font(name=XLS_FONT, size=10, bold=True, color=XLS_ROJO_HEADER)
    ws.row_dimensions[fila].height = 18
    return fila + 1


def _xls_tabla(ws, fila, headers, filas, col_widths, currency_cols=None, right_cols=None, color_col=None):
    """
    Dibuja una tabla estilo CYS: cabecera roja, filas alternas rosa suave/blanco,
    bordes suaves. color_col = {col_index_0based: {'valor_texto': color_hex}} opcional
    para colorear una celda según su contenido (ej. Estado).
    Devuelve la fila siguiente disponible.
    """
    fills = _xls_fills()
    borde = _xls_borde_suave()
    currency_cols = currency_cols or set()
    right_cols = right_cols or set()
    color_col = color_col or {}

    fila_header = fila
    for j, texto in enumerate(headers, start=1):
        c = ws.cell(row=fila_header, column=j, value=texto)
        c.font = Font(name=XLS_FONT, size=10, bold=True, color=XLS_BLANCO)
        c.fill = fills['header']
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = borde
    ws.row_dimensions[fila_header].height = 20

    for i, datos_fila in enumerate(filas):
        r = fila_header + 1 + i
        fondo = fills['suave'] if i % 2 == 1 else fills['blanco']
        for j, valor in enumerate(datos_fila, start=1):
            c = ws.cell(row=r, column=j, value=valor)
            c.fill = fondo
            c.border = borde
            c.font = Font(name=XLS_FONT, size=9, color=XLS_GRIS_TEXTO)
            if (j - 1) in currency_cols and isinstance(valor, (int, float)):
                c.number_format = '"$"#,##0'
                c.alignment = Alignment(horizontal='right')
            elif (j - 1) in right_cols:
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center')
            if (j - 1) in color_col and str(valor) in color_col[j - 1]:
                c.font = Font(name=XLS_FONT, size=9, bold=True, color=color_col[j - 1][str(valor)])
        ws.row_dimensions[r].height = 16

    for j, ancho in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ancho

    ultima_fila = fila_header + len(filas)
    ws.freeze_panes = ws.cell(row=fila_header + 1, column=1)
    if filas:
        ws.auto_filter.ref = f"A{fila_header}:{get_column_letter(len(headers))}{ultima_fila}"

    return ultima_fila + 2  # deja una fila en blanco después de la tabla


def _xlsx_ventas(wb, ventas_qs, fecha_inicio, fecha_fin):
    ws = wb.active
    ws.title = 'Historial de Ventas'
    ws.sheet_view.showGridLines = False

    total_v = sum(v.total_venta for v in ventas_qs)
    total_u = sum(det.cantidad for v in ventas_qs for det in v.detalles.all())
    total_c = ventas_qs.values('cliente').distinct().count()

    subtitulo = f"Período: {fecha_inicio or 'Todo'} → {fecha_fin or 'Todo'}"
    meta = [('Generado:', timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M'))]
    fila = _xls_encabezado(ws, "Historial de Ventas", subtitulo, meta, 7)
    fila = _xls_kpis(ws, fila, [
        ('Total Ventas', total_v, XLS_VERDE_OK),
        ('Unidades', total_u, XLS_ROJO_HEADER),
        ('Clientes únicos', total_c, XLS_ROJO_MARCA),
    ], 7)

    headers = ['#', 'Fecha', 'Cliente', 'Producto', 'Cantidad', 'Precio Unit.', 'Total']
    filas = []
    contador = 1
    for v in ventas_qs:
        for det in v.detalles.all():
            filas.append([
                contador,
                v.fecha.astimezone(ZONA_COLOMBIA).strftime("%Y-%m-%d"),
                str(v.cliente),
                det.producto.nombre,
                det.cantidad,
                float(det.precio_unitario),
                float(det.subtotal()),
            ])
            contador += 1

    _xls_tabla(ws, fila, headers, filas,
               col_widths=[6, 13, 24, 32, 11, 15, 15],
               currency_cols={5, 6}, right_cols={4})


def _xlsx_inventario(wb, productos_qs, entradas_qs, salidas_qs):
    ws = wb.active
    ws.title = 'Inventario'
    ws.sheet_view.showGridLines = False

    total_reg  = productos_qs.count()
    en_stock   = sum(1 for p in productos_qs if p.stock_total > 10)
    stock_bajo = sum(1 for p in productos_qs if 0 < p.stock_total <= 10)
    agotados   = sum(1 for p in productos_qs if p.stock_total == 0)

    meta = [('Generado:', timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M'))]
    fila = _xls_encabezado(ws, "Reporte de Inventario", "Stock, niveles y movimientos", meta, 5)
    fila = _xls_kpis(ws, fila, [
        ('Registrados', total_reg, XLS_ROJO_MARCA),
        ('En stock', en_stock, XLS_VERDE_OK),
        ('Stock bajo', stock_bajo, 'B8860B'),
        ('Agotados', agotados, XLS_ROJO_OUT),
    ], 5)

    fila = _xls_seccion(ws, fila, 'Estado de Stock', 5)
    headers = ['#', 'Producto', 'Stock', 'Estado']
    filas = []
    for i, p in enumerate(productos_qs, 1):
        qty = p.stock_total
        estado = 'Agotado' if qty == 0 else ('Stock bajo' if qty <= 10 else 'En stock')
        filas.append([i, p.nombre, qty, estado])
    fila = _xls_tabla(
        ws, fila, headers, filas,
        col_widths=[6, 42, 12, 16], right_cols={2},
        color_col={3: {'Agotado': XLS_ROJO_OUT, 'Stock bajo': 'B8860B', 'En stock': XLS_VERDE_OK}},
    )

    fila = _xls_seccion(ws, fila, 'Entradas de Inventario', 5)
    headers_e = ['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']
    filas_e = []
    for i, e in enumerate(entradas_qs, 1):
        filas_e.append([
            i, e.presentacion.producto.nombre, f'+{e.cantidad}',
            e.motivo or '—',
            e.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M"),
        ])
    fila = _xls_tabla(ws, fila, headers_e, filas_e, col_widths=[6, 34, 12, 22, 18], right_cols={2})

    fila = _xls_seccion(ws, fila, 'Salidas de Inventario', 5)
    headers_s = ['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']
    filas_s = []
    for i, s in enumerate(salidas_qs, 1):
        filas_s.append([
            i, s.presentacion.producto.nombre, f'-{s.cantidad}',
            s.motivo or '—',
            s.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M"),
        ])
    _xls_tabla(ws, fila, headers_s, filas_s, col_widths=[6, 34, 12, 22, 18], right_cols={2})


def _xlsx_proveedores(wb, proveedores_list):
    ws = wb.active
    ws.title = 'Proveedores'
    ws.sheet_view.showGridLines = False

    meta = [('Generado:', timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M'))]
    fila = _xls_encabezado(ws, "Reporte de Proveedores", "Lista y datos de proveedores activos", meta, 5)
    fila = _xls_kpis(ws, fila, [
        ('Total Proveedores', len(proveedores_list), XLS_ROJO_MARCA),
        ('Estado', 'Activos', XLS_VERDE_OK),
    ], 5)

    headers = ['#', 'Contacto', 'Empresa', 'Teléfono', 'Correo']
    filas = []
    for i, p in enumerate(proveedores_list, 1):
        filas.append([i, p['nombre_contacto'], p['nombre_empresa'], p['telefono'], p['email']])
    _xls_tabla(ws, fila, headers, filas, col_widths=[6, 22, 30, 16, 30])


def _xlsx_resumen_diario(wb, hoy, ventas_hoy, entradas_hoy, salidas_hoy,
                          ingresos_hoy, total_entradas_hoy, total_salidas_hoy,
                          top_productos_hoy):
    ws = wb.active
    ws.title = 'Resumen Diario'
    ws.sheet_view.showGridLines = False

    meta = [('Fecha:', hoy.strftime('%d/%m/%Y')),
            ('Generado:', timezone.now().astimezone(ZONA_COLOMBIA).strftime('%H:%M'))]
    fila = _xls_encabezado(ws, "Resumen Diario", f"CYS Ltda. · {hoy.strftime('%d/%m/%Y')}", meta, 6)
    fila = _xls_kpis(ws, fila, [
        ('Ingresos hoy', ingresos_hoy, XLS_VERDE_OK),
        ('Ventas hoy', len(ventas_hoy), XLS_ROJO_MARCA),
        ('Und. Entrada', total_entradas_hoy, XLS_VERDE_OK),
        ('Und. Salida', total_salidas_hoy, XLS_ROJO_OUT),
    ], 6)

    fila = _xls_seccion(ws, fila, 'Ventas del día', 6)
    headers_v = ['#', 'Cliente', 'Producto', 'Cantidad', 'Total', 'Hora']
    filas_v = []
    cont = 1
    for v in ventas_hoy:
        for det in v.detalles.all():
            filas_v.append([
                cont, str(v.cliente), det.producto.nombre, det.cantidad,
                float(det.subtotal()), v.fecha.astimezone(ZONA_COLOMBIA).strftime("%H:%M"),
            ])
            cont += 1
    fila = _xls_tabla(ws, fila, headers_v, filas_v,
                       col_widths=[6, 22, 30, 11, 14, 10], currency_cols={4}, right_cols={3})

    fila = _xls_seccion(ws, fila, 'Entradas del día', 6)
    headers_e = ['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']
    filas_e = []
    for i, e in enumerate(entradas_hoy, 1):
        filas_e.append([i, e.presentacion.producto.nombre, f'+{e.cantidad}',
                         e.motivo or '—',
                         e.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M")])
    fila = _xls_tabla(ws, fila, headers_e, filas_e, col_widths=[6, 30, 12, 22, 18], right_cols={2})

    fila = _xls_seccion(ws, fila, 'Salidas del día', 6)
    headers_s = ['#', 'Producto', 'Cantidad', 'Motivo', 'Fecha']
    filas_s = []
    for i, s in enumerate(salidas_hoy, 1):
        filas_s.append([i, s.presentacion.producto.nombre, f'-{s.cantidad}',
                         s.motivo or '—',
                         s.fecha_actualizada.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y %H:%M")])
    fila = _xls_tabla(ws, fila, headers_s, filas_s, col_widths=[6, 30, 12, 22, 18], right_cols={2})

    if top_productos_hoy:
        fila = _xls_seccion(ws, fila, 'Top productos del día', 6)
        headers_t = ['#', 'Producto', 'Unidades vendidas', 'Total generado']
        filas_t = []
        for i, (nombre, datos) in enumerate(top_productos_hoy, 1):
            filas_t.append([i, nombre, datos['cantidad'], float(datos['subtotal'])])
        _xls_tabla(ws, fila, headers_t, filas_t, col_widths=[6, 32, 18, 18],
                   currency_cols={3}, right_cols={2})


def _xlsx_analisis_ventas(wb, ventas_qs, total_ventas, total_productos, total_clientes, top_productos_hoy):
    ws = wb.active
    ws.title = 'Análisis de Ventas'
    ws.sheet_view.showGridLines = False

    meta = [('Generado:', timezone.now().astimezone(ZONA_COLOMBIA).strftime('%d/%m/%Y %H:%M'))]
    fila = _xls_encabezado(ws, "Análisis de Ventas", "Resumen de ventas y top productos", meta, 5)
    fila = _xls_kpis(ws, fila, [
        ('Ventas Totales', total_ventas, XLS_VERDE_OK),
        ('Uds. Vendidas', total_productos, XLS_ROJO_MARCA),
        ('Clientes', total_clientes, 'B8860B'),
    ], 5)

    fila = _xls_seccion(ws, fila, 'Ventas por fecha', 5)
    headers = ['Fecha', 'Cliente', 'Producto', 'Cantidad', 'Total']
    filas = []
    for v in ventas_qs:
        for det in v.detalles.all():
            filas.append([
                v.fecha.astimezone(ZONA_COLOMBIA).strftime("%d/%m/%Y"),
                str(v.cliente), det.producto.nombre, det.cantidad, float(det.subtotal()),
            ])
    fila = _xls_tabla(ws, fila, headers, filas,
                       col_widths=[14, 22, 32, 12, 15], currency_cols={4}, right_cols={3})

    fila = _xls_seccion(ws, fila, 'Top productos del día más vendidos', 5)
    headers_t = ['#', 'Producto', 'Unidades vendidas', 'Total generado']
    filas_t = []
    for i, (nombre, datos) in enumerate(top_productos_hoy, 1):
        filas_t.append([i, nombre, datos['cantidad'], float(datos['subtotal'])])
    _xls_tabla(ws, fila, headers_t, filas_t, col_widths=[6, 32, 18, 18],
               currency_cols={3}, right_cols={2})


# ═══════════════════════════════════════════════════════
#  VISTA PRINCIPAL
# ═══════════════════════════════════════════════════════

def index_reportes(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin    = request.GET.get('fecha_fin')
    per_page     = request.GET.get('per_page', '10')

    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    ventas_qs = Venta.objects.prefetch_related(
        'detalles__producto',
        'detalles__presentacion'
    ).all().order_by('-fecha')

    if fecha_inicio:
        ventas_qs = ventas_qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        ventas_qs = ventas_qs.filter(fecha__date__lte=fecha_fin)

    paginator   = Paginator(ventas_qs, per_page)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    proveedores_data_list = [
        {"nombre_contacto": "Carlos Mendoza",      "nombre_empresa": "Licores del Caribe S.A.S.",    "telefono": "300 456 7890", "email": "carlos.mendoza@licorescaribe.com"},
        {"nombre_contacto": "Ana María Gómez",     "nombre_empresa": "Distribuidora El Brindis",      "telefono": "315 987 6543", "email": "contacto@elbrindis.com"},
        {"nombre_contacto": "Juan Fernando Ruiz",  "nombre_empresa": "Cervecerías Unidas",            "telefono": "310 456 1234", "email": "jruiz@cerveceriasunidas.co"},
        {"nombre_contacto": "Patricia Lara",        "nombre_empresa": "Importaciones Premium Ltda.",  "telefono": "312 789 4561", "email": "plara@premiumimports.com"},
        {"nombre_contacto": "Roberto Díaz",         "nombre_empresa": "Bodegas de la Sabana",         "telefono": "320 123 7894", "email": "rdiaz@bodegassabana.com"},
    ]

    # ─────────────────────────────────────────────────────
    #  EXPORTACIONES
    # ─────────────────────────────────────────────────────
    export_format = request.GET.get('export')

    if export_format in ('excel', 'pdf'):
        tipo = request.GET.get('tipo', 'ventas')

        # ── EXCEL (.xlsx real, con estilos) ───────────────
        if export_format == 'excel':
            productos_qs = Producto.objects.all().order_by('nombre')
            hoy_xls      = timezone.now().astimezone(ZONA_COLOMBIA).date()

            wb = Workbook()

            if tipo == 'ventas':
                _xlsx_ventas(wb, ventas_qs, fecha_inicio, fecha_fin)
                nombre = f"reporte_ventas_{timezone.now().strftime('%Y%m%d')}.xlsx"

            elif tipo == 'inventario':
                entradas_qs = Inventario.objects.filter(tipo='entrada').select_related('presentacion__producto').order_by('-fecha_actualizada')
                salidas_qs  = Inventario.objects.filter(tipo='salida').select_related('presentacion__producto').order_by('-fecha_actualizada')
                _xlsx_inventario(wb, productos_qs, entradas_qs, salidas_qs)
                nombre = f"reporte_inventario_{timezone.now().strftime('%Y%m%d')}.xlsx"

            elif tipo == 'proveedores':
                _xlsx_proveedores(wb, proveedores_data_list)
                nombre = f"reporte_proveedores_{timezone.now().strftime('%Y%m%d')}.xlsx"

            elif tipo == 'resumen_diario':
                ventas_hoy_xls   = Venta.objects.prefetch_related('detalles__producto', 'detalles__presentacion').filter(fecha__date=hoy_xls).order_by('-fecha')
                ingresos_hoy_xls = sum(v.total_venta for v in ventas_hoy_xls)
                mov_hoy          = Inventario.objects.filter(fecha_actualizada__date=hoy_xls).select_related('presentacion__producto')
                entradas_hoy_xls = mov_hoy.filter(tipo='entrada')
                salidas_hoy_xls  = mov_hoy.filter(tipo='salida')
                total_ent_xls    = sum(e.cantidad for e in entradas_hoy_xls)
                total_sal_xls    = sum(s.cantidad for s in salidas_hoy_xls)

                det_hoy = DetalleVenta.objects.filter(venta__fecha__date=hoy_xls).select_related('producto')
                top_h = {}
                for det in det_hoy:
                    n = det.producto.nombre
                    if n not in top_h:
                        top_h[n] = {'cantidad': 0, 'subtotal': 0}
                    top_h[n]['cantidad'] += det.cantidad
                    top_h[n]['subtotal'] += float(det.subtotal())
                top_h = sorted(top_h.items(), key=lambda x: x[1]['subtotal'], reverse=True)[:5]

                _xlsx_resumen_diario(
                    wb, hoy_xls, ventas_hoy_xls, entradas_hoy_xls, salidas_hoy_xls,
                    ingresos_hoy_xls, total_ent_xls, total_sal_xls, top_h
                )
                nombre = f"resumen_diario_{hoy_xls.strftime('%Y%m%d')}.xlsx"

            elif tipo == 'analisis':
                total_v = sum(v.total_venta for v in ventas_qs)
                total_u = sum(det.cantidad for v in ventas_qs for det in v.detalles.all())
                total_c = ventas_qs.values('cliente').distinct().count()

                det_qs = DetalleVenta.objects.filter(venta__fecha__date=hoy_xls).select_related('producto')
                top_h = {}
                for det in det_qs:
                    n = det.producto.nombre
                    if n not in top_h:
                        top_h[n] = {'cantidad': 0, 'subtotal': 0}
                    top_h[n]['cantidad'] += det.cantidad
                    top_h[n]['subtotal'] += float(det.subtotal())
                top_h = sorted(top_h.items(), key=lambda x: x[1]['subtotal'], reverse=True)[:5]

                _xlsx_analisis_ventas(wb, ventas_qs, total_v, total_u, total_c, top_h)
                nombre = f"analisis_ventas_{timezone.now().strftime('%Y%m%d')}.xlsx"

            else:
                return HttpResponse("Tipo de reporte Excel no reconocido.", status=400)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'
            return response

        # ── PDF ──────────────────────────────────────────
        elif export_format == 'pdf':
            productos_qs = Producto.objects.all().order_by('nombre')
            hoy_pdf      = timezone.now().astimezone(ZONA_COLOMBIA).date()

            if tipo == 'ventas':
                buffer = _pdf_ventas(ventas_qs, fecha_inicio, fecha_fin)
                nombre = f"reporte_ventas_{timezone.now().strftime('%Y%m%d')}.pdf"

            elif tipo == 'inventario':
                entradas_qs = Inventario.objects.filter(tipo='entrada').select_related('presentacion__producto').order_by('-fecha_actualizada')
                salidas_qs  = Inventario.objects.filter(tipo='salida').select_related('presentacion__producto').order_by('-fecha_actualizada')
                buffer = _pdf_inventario(productos_qs, entradas_qs, salidas_qs)
                nombre = f"reporte_inventario_{timezone.now().strftime('%Y%m%d')}.pdf"

            elif tipo == 'proveedores':
                buffer = _pdf_proveedores(proveedores_data_list)
                nombre = f"reporte_proveedores_{timezone.now().strftime('%Y%m%d')}.pdf"

            elif tipo == 'resumen_diario':
                ventas_hoy_pdf    = Venta.objects.prefetch_related('detalles__producto', 'detalles__presentacion').filter(fecha__date=hoy_pdf).order_by('-fecha')
                ingresos_hoy_pdf  = sum(v.total_venta for v in ventas_hoy_pdf)
                mov_hoy           = Inventario.objects.filter(fecha_actualizada__date=hoy_pdf).select_related('presentacion__producto')
                entradas_hoy_pdf  = mov_hoy.filter(tipo='entrada')
                salidas_hoy_pdf   = mov_hoy.filter(tipo='salida')
                total_ent_pdf     = sum(e.cantidad for e in entradas_hoy_pdf)
                total_sal_pdf     = sum(s.cantidad for s in salidas_hoy_pdf)

                det_hoy = DetalleVenta.objects.filter(venta__fecha__date=hoy_pdf).select_related('producto')
                top_h   = {}
                for det in det_hoy:
                    n = det.producto.nombre
                    if n not in top_h:
                        top_h[n] = {'cantidad': 0, 'subtotal': 0}
                    top_h[n]['cantidad'] += det.cantidad
                    top_h[n]['subtotal'] += float(det.subtotal())
                top_h = sorted(top_h.items(), key=lambda x: x[1]['subtotal'], reverse=True)[:5]

                buffer = _pdf_resumen_diario(
                    hoy_pdf, ventas_hoy_pdf, entradas_hoy_pdf, salidas_hoy_pdf,
                    ingresos_hoy_pdf, total_ent_pdf, total_sal_pdf, top_h
                )
                nombre = f"resumen_diario_{hoy_pdf.strftime('%Y%m%d')}.pdf"

            elif tipo == 'analisis':
                total_v = sum(v.total_venta for v in ventas_qs)
                total_u = sum(det.cantidad for v in ventas_qs for det in v.detalles.all())
                total_c = ventas_qs.values('cliente').distinct().count()

                det_qs = DetalleVenta.objects.filter(venta__fecha__date=hoy_pdf).select_related('producto')
                top_h  = {}
                for det in det_qs:
                    n = det.producto.nombre
                    if n not in top_h:
                        top_h[n] = {'cantidad': 0, 'subtotal': 0}
                    top_h[n]['cantidad'] += det.cantidad
                    top_h[n]['subtotal'] += float(det.subtotal())
                top_h = sorted(top_h.items(), key=lambda x: x[1]['subtotal'], reverse=True)[:5]

                buffer = _pdf_analisis_ventas(ventas_qs, total_v, total_u, total_c, top_h)
                nombre = f"analisis_ventas_{timezone.now().strftime('%Y%m%d')}.pdf"

            else:
                return HttpResponse("Tipo de reporte PDF no reconocido.", status=400)

            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{nombre}"'
            return response

    # ─────────────────────────────────────────────────────
    #  RENDER NORMAL
    # ─────────────────────────────────────────────────────
    ventas_todas = ventas_qs

    productos   = Producto.objects.all().order_by('nombre')
    proveedores = proveedores_data_list

    total_ventas    = sum(v.total_venta for v in ventas_todas)
    total_productos = sum(det.cantidad for v in ventas_todas for det in v.detalles.all())
    total_clientes  = ventas_todas.values('cliente').distinct().count()

    total_registrados = productos.count()
    total_en_stock    = sum(1 for p in productos if p.stock_total > 10)
    total_stock_bajo  = sum(1 for p in productos if 0 < p.stock_total <= 10)
    total_agotados    = sum(1 for p in productos if p.stock_total == 0)

    entradas = Inventario.objects.filter(tipo='entrada').select_related('presentacion__producto').order_by('-fecha_actualizada')
    salidas  = Inventario.objects.filter(tipo='salida').select_related('presentacion__producto').order_by('-fecha_actualizada')

    hoy = timezone.now().astimezone(ZONA_COLOMBIA).date()

    ventas_hoy = Venta.objects.prefetch_related(
        'detalles__producto', 'detalles__presentacion'
    ).filter(fecha__date=hoy).order_by('-fecha')

    ingresos_hoy = sum(v.total_venta for v in ventas_hoy)

    movimientos_hoy    = Inventario.objects.filter(fecha_actualizada__date=hoy).select_related('presentacion__producto')
    entradas_hoy       = movimientos_hoy.filter(tipo='entrada')
    salidas_hoy        = movimientos_hoy.filter(tipo='salida')
    total_entradas_hoy = sum(e.cantidad for e in entradas_hoy)
    total_salidas_hoy  = sum(s.cantidad for s in salidas_hoy)

    productos_vendidos_hoy = (
        DetalleVenta.objects
        .filter(venta__fecha__date=hoy)
        .select_related('producto', 'presentacion', 'venta')
        .order_by('producto__nombre')
    )
    top_productos_hoy = {}
    for det in productos_vendidos_hoy:
        nombre = det.producto.nombre
        if nombre not in top_productos_hoy:
            top_productos_hoy[nombre] = {'cantidad': 0, 'subtotal': 0}
        top_productos_hoy[nombre]['cantidad'] += det.cantidad
        top_productos_hoy[nombre]['subtotal'] += float(det.subtotal())
    top_productos_hoy = sorted(
        top_productos_hoy.items(),
        key=lambda x: x[1]['subtotal'],
        reverse=True
    )[:5]

    ventas_data = []
    for v in ventas_todas:
        for det in v.detalles.all():
            ventas_data.append({
                "fecha":           v.fecha.astimezone(ZONA_COLOMBIA).strftime("%Y-%m-%d"),
                "hora":            v.fecha.astimezone(ZONA_COLOMBIA).strftime("%H:%M"),
                "cliente":         str(v.cliente),
                "producto":        det.producto.nombre,
                "presentacion":    det.presentacion.nombre if det.presentacion else "Unidad",
                "cantidad":        det.cantidad,
                "precio_unitario": float(det.precio_unitario),
                "subtotal":        float(det.subtotal()),
                "descuento":       float(v.descuento_porcentaje),
                "total_venta":     float(v.total_venta),
            })

    ventas_json = (
        json.dumps(ventas_data, cls=DjangoJSONEncoder)
        .replace('</script>', r'<\/script>')
        .replace('<!--',      r'<\!--')
    )

    context = {
        'ventas':             ventas_todas,
        'page_obj':           page_obj,
        'paginator':          paginator,
        'total_ventas':       total_ventas,
        'total_productos':    total_productos,
        'total_clientes':     total_clientes,
        'productos':          productos,
        'proveedores':        proveedores,
        'total_registrados':  total_registrados,
        'total_en_stock':     total_en_stock,
        'total_stock_bajo':   total_stock_bajo,
        'total_agotados':     total_agotados,
        'entradas':           entradas,
        'salidas':            salidas,
        'fecha_inicio':       fecha_inicio or '',
        'fecha_fin':          fecha_fin or '',
        'per_page':           per_page,
        'hoy':                hoy,
        'ventas_hoy':         ventas_hoy,
        'ingresos_hoy':       ingresos_hoy,
        'entradas_hoy':       entradas_hoy,
        'salidas_hoy':        salidas_hoy,
        'total_entradas_hoy': total_entradas_hoy,
        'total_salidas_hoy':  total_salidas_hoy,
        'top_productos_hoy':  top_productos_hoy,
        'ventas_json':        ventas_json,
        'breadcrumb_items': [
            {'nombre': 'Reportes', 'url': None},
        ],
    }
    return render(request, 'reportes/reportes.html', context)


# ═══════════════════════════════════════════════════════
#  VISTA MOVIMIENTOS
# ═══════════════════════════════════════════════════════

def reporte_movimientos(request):
    form        = FiltroReporteForm(request.GET or None)
    movimientos = Inventario.objects.select_related('presentacion__producto').all().order_by('-fecha_actualizada')

    if form.is_valid():
        f_inicio = form.cleaned_data.get('fecha_inicio')
        f_fin    = form.cleaned_data.get('fecha_fin')
        tipo     = form.cleaned_data.get('tipo_reporte')

        if f_inicio:
            movimientos = movimientos.filter(fecha_actualizada__date__gte=f_inicio)
        if f_fin:
            movimientos = movimientos.filter(fecha_actualizada__date__lte=f_fin)
        if tipo and tipo != 'general':
            movimientos = movimientos.filter(tipo=tipo)

    paginator = Paginator(movimientos, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'form':        form,
        'movimientos': page_obj,
        'breadcrumb_items': [
            {'nombre': 'Reportes', 'url': reverse('index_reportes')},
            {'nombre': 'Movimientos', 'url': None},
        ],
    }
    return render(request, 'reportes/reporte_movimientos.html', context)